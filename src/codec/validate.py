"""
validate.py — compute projection quality heuristics and export to Excel.

Reads a JSONL file produced by :mod:`src.codec.translate` (which contains
projected span annotations) and computes a battery of heuristic flags that
indicate whether the projections look plausible:

* Bracket-count mismatches between source and target tagged texts
* Whether projected spans split words in the target language
* Whether projected spans end on punctuation
* Whether any projected span extends beyond the sentence boundary

Additionally, it translates each projected span independently (without
sentence context) using NLLB, computes pairwise BERTScore F1 similarity
between original and projected spans via the Hungarian algorithm, and writes
everything to a formatted Excel workbook.

Usage
-----
.. code-block:: bash

    python src/codec/validate.py \\
        --input_path  tests/output2.jsonl \\
        --output_path tests/validation.xlsx \\
        --model_name_or_path XIE2021/nllb-200-3.3B-easyproject \\
        --tokenizer_path facebook/nllb-200-distilled-600M \\
        --src_lang en \\
        --tgt_lang de

All ``--model_*`` arguments are forwarded to :class:`transformers.AutoModelForSeq2SeqLM`.
"""
from __future__ import annotations

import argparse
import json
import os
import re
import string
from enum import Enum
from timeit import default_timer as timer
from typing import Dict, List, Tuple

import pandas as pd
import regex
from scipy.optimize import linear_sum_assignment

# Silence HF / transformers logs and warnings before any import that triggers them.
import logging
import warnings

logging.getLogger("huggingface_hub").setLevel(logging.ERROR)
logging.getLogger("transformers").setLevel(logging.ERROR)
logging.getLogger("torch").setLevel(logging.WARNING)

# Suppress all FutureWarnings (e.g., from huggingface_hub about deprecated args)
warnings.filterwarnings("ignore", category=FutureWarning, module="huggingface_hub")

os.environ["TOKENIZERS_PARALLELISM"] = "false"
os.environ["TRANSFORMERS_NO_ADVISORY_WARNINGS"] = "true"
os.environ["HF_HUB_DISABLE_TELEMETRY"] = "1"
os.environ["HF_HUB_DISABLE_PROGRESS_BARS"] = "1"
os.environ["HF_HUB_VERBOSITY"] = "critical"

import torch
from bert_score import BERTScorer
from openpyxl.styles import Alignment, Font
from transformers import AutoModelForSeq2SeqLM, AutoTokenizer
from transformers.utils.logging import disable_progress_bar, set_verbosity_error
disable_progress_bar()
set_verbosity_error()


logger = logging.getLogger(__name__)

from src.codec.translate import _iso2_to_nllb, labels_to_bracketed


# ---------------------------------------------------------------------------
# Human-readable labels for each heuristic
# ---------------------------------------------------------------------------

class IssueTypes(Enum):
    __order__ = (
        "n_opening_mismatch n_closing_mismatch "
        "src_span_splits_word src_span_ends_with_punctuation "
        "src_span_extends_whole_sentence"
    )
    n_opening_mismatch = "Mismatching opening bracket(s)"
    n_closing_mismatch = "Mismatching closing bracket(s)"
    src_span_splits_word = "Projected label splits word"
    src_span_ends_with_punctuation = "Projected label ends with punctuation"
    src_span_extends_whole_sentence = "whole-sentence span"


# ---------------------------------------------------------------------------
# Heuristic helpers
# ---------------------------------------------------------------------------

def _count_brackets(x: pd.Series | list[str] | str, char: str) -> int | pd.Series:
    """Count occurrences of ``char`` in each string."""
    if isinstance(x, str):
        return x.count(char)
    if isinstance(x, list):
        x = pd.Series(x)
    return x.str.count(re.escape(char)).fillna(0).astype(int)


def _span_splits_word(text: str, start: int, end: int) -> bool:
    """Check if span ``[start:end]`` cuts through a word boundary."""
    if start > 0 and not text[start - 1].isspace():
        return True
    if end < len(text) and not (text[end].isspace() or text[end] in string.punctuation):
        return True
    return False


def _span_ends_with_punctuation(text: str, start: int, end: int) -> bool:
    """Check if span ``[start:end]`` ends with punctuation."""
    return end > 0 and text[end - 1] in string.punctuation


def _span_extends_sentence(text: str, start: int, end: int) -> bool:
    """Check if span ``[start:end]`` exceeds the source text length."""
    return start == 0 and end >= len(text)


def _create_issues_list(out: pd.DataFrame) -> list[str]:
    return [
        ", ".join(
            issue.value for issue in IssueTypes if row[issue.name] == 1
        )
        for _, row in out.iterrows()
    ]


# ---------------------------------------------------------------------------
# Core pipeline functions
# ---------------------------------------------------------------------------

def compute_projection_heuristics(df: pd.DataFrame) -> pd.DataFrame:
    """Compute projection quality indicators on a batch DataFrame.

    Parameters
    ----------
    df : pd.DataFrame
        Must contain columns: ``text``, ``source_text``, ``label``,
        ``source_label``.

    Returns
    -------
    pd.DataFrame
        Original DataFrame with heuristic indicator columns appended.
    """
    out = df.copy()

    # remove any control characters in unicode range 0x01 - 0x02
    out['text'] = out['text'].apply(lambda x: regex.sub(r'[\x01-\x02]', '', x))
    out['source_text'] = out['source_text'].apply(lambda x: regex.sub(r'[\x01-\x02]', '', x))

    # replace any brackets in the source text with a placeholder before tagging, then restore them afterward
    out["text"] = out["text"].str.replace("[", "\x01", regex=False)
    out["text"] = out["text"].str.replace("]", "\x02", regex=False)
    out["source_text"] = out["source_text"].str.replace("[", "\x01", regex=False)
    out["source_text"] = out["source_text"].str.replace("]", "\x02", regex=False)

    out['text_tagged'] = out.apply(lambda row: labels_to_bracketed(row.text, row.label)[0], axis=1)
    out['source_text_tagged'] = out.apply(lambda row: labels_to_bracketed(row.source_text, row.source_label)[0], axis=1)

    # --- Bracket count mismatches (coarse-grained) ---
    out["n_opening_brackets_en"]  = n_open_en  = _count_brackets(out["text_tagged"], "[")
    out["n_opening_brackets_src"] = n_open_src = _count_brackets(out["source_text_tagged"], "[")
    out["n_opening_mismatch"] = (n_open_en != n_open_src).astype(int)

    out["n_closing_brackets_en"]  = n_close_en  = _count_brackets(out["text_tagged"], "]")
    out["n_closing_brackets_src"] = n_close_src = _count_brackets(out["source_text_tagged"], "]")
    out["n_closing_mismatch"] = (n_close_en != n_close_src).astype(int)

    # --- Per-span checks (fine-grained, applied to source_label) ---
    def _check_splits(row):
        text = row["source_text"]
        labels = row["source_label"]
        if not labels:
            return False
        return any(_span_splits_word(text, s, e) for s, e, *_ in labels)

    def _check_ends_with_punctuation(row):
        text = row["source_text"]
        labels = row["source_label"]
        if not labels:
            return False
        return any(_span_ends_with_punctuation(text, s, e) for s, e, *_ in labels)

    def _check_extends(row):
        text = row["source_text"]
        labels = row["source_label"]
        if not labels:
            return False
        return any(_span_extends_sentence(text, s, e) for s, e, *_ in labels)

    out["src_span_splits_word"] = out.apply(_check_splits, axis=1).astype(int)
    out["src_span_ends_with_punctuation"] = out.apply(_check_ends_with_punctuation, axis=1).astype(int)
    out["src_span_extends_whole_sentence"] = out.apply(_check_extends, axis=1).astype(int)

    out["n_issues"] = out[IssueTypes._member_names_].sum(axis=1)
    out["issues"] = _create_issues_list(out)

    return out


def extract_spans(text: str, label: list[list]) -> list[str]:
    """Extract span texts from a doccano-style annotation list."""
    return [text[s:e] for s, e, *_ in label]


def compute_spans_word_counts(
    text: str, label: list[list], tokenizer, as_dict: bool = False
) -> list[int] | dict[str, int]:
    """Compute word (token-set) counts for each span."""
    spans = extract_spans(text, label)
    toks = tokenizer(spans, add_special_tokens=False)
    word_counts = [len(set(toks.word_ids(i))) for i in range(len(spans))]
    if as_dict:
        return dict(zip(spans, word_counts))
    return word_counts


def translate_spans(
    text: str, label: list[list], lang: str, model, tokenizer,
    max_length: int = 48, as_dict: bool = False,
) -> list[str] | dict[str, str]:
    """Translate each span independently using NLLB (no sentence context)."""
    if not label or len(label) == 0:
        return [] if not as_dict else {}

    spans = extract_spans(text, label)
    tokenizer.src_lang = _iso2_to_nllb(lang)
    inputs = tokenizer(
        spans, padding=True, truncation=True,
        max_length=max_length, return_tensors="pt",
    ).to(model.device)

    with torch.no_grad():
        translated_tokens = model.generate(
            **inputs,
            forced_bos_token_id=tokenizer.lang_code_to_id["eng_Latn"],
            num_beams=5,
            max_length=max_length,
            num_return_sequences=1,
            early_stopping=True,
        )

    translated = tokenizer.batch_decode(translated_tokens, skip_special_tokens=True)

    if as_dict:
        return dict(zip(spans, translated))
    return translated


# ---------------------------------------------------------------------------
# BERTScore matcher
# ---------------------------------------------------------------------------

class BERTScoreMatcher:
    """Compute pairwise BERTScore F1 between source and projected spans."""

    def __init__(self, model_name: str):
        self.scorer = BERTScorer(
            model_type=model_name,
            device="cuda" if torch.cuda.is_available() else "cpu",
        )

    def __call__(
        self, source_spans: list[str], translated_spans: list[str]
    ) -> pd.DataFrame:
        """Compute pairwise BERTScore F1 and return a DataFrame."""
        if not source_spans or not translated_spans:
            return pd.DataFrame(
                columns=["source_span", "translated_span", "score"]
            )

        a = [s for s in source_spans for _ in translated_spans]
        b = [t for _ in source_spans for t in translated_spans]

        # Compute pairwise BERTScore F1
        *_, f1s = self.scorer.score(a, b)

        # create a 2D array of F1 scores
        sims = f1s.reshape((len(source_spans), len(translated_spans)))

        # apply Hungarian algorithm (minimizes total cost = maximizes total F1)
        ridxs, cidxs = linear_sum_assignment(1 - sims)

        matched = pd.DataFrame([
            [source_spans[r], translated_spans[c], sims[r, c].item()]
            for r, c in zip(ridxs, cidxs)
        ])
        matched.columns = ["source_span", "translated_span", "score"]
        matched["score"] = matched["score"].astype(float).round(4)

        return matched


# ---------------------------------------------------------------------------
# Excel export helpers
# ---------------------------------------------------------------------------

class FirstLevelIndentEncoder(json.JSONEncoder):
    """Pretty-print JSON with indented top-level containers."""

    def encode(self, obj):
        if isinstance(obj, list):
            lines = [
                f"  {json.dumps(item, ensure_ascii=self.ensure_ascii)}"
                for item in obj
            ]
            return "[\n" + ",\n".join(lines) + "\n]"
        if isinstance(obj, dict):
            lines = [
                f"  {json.dumps({k: v}, ensure_ascii=self.ensure_ascii)}"
                for k, v in obj.items()
            ]
            return "{\n" + ",\n".join(lines) + "\n}"
        return super().encode(obj)


_prettify = lambda x: json.dumps(x, cls=FirstLevelIndentEncoder, ensure_ascii=False)

_HIDDEN_COLS = [
    "metadata",
    "text",
    "source_text",
    "label",
    "source_label",
    "n_opening_brackets_en",
    "n_opening_brackets_src",
    "n_closing_brackets_en",
    "n_closing_brackets_src",
    "n_opening_mismatch",
    "n_closing_mismatch",
    "src_span_splits_word",
    "src_span_ends_with_punctuation",
    "src_span_extends_whole_sentence",
]

_WIDE_COLS = {
    "text_tagged": 60,
    "source_text_tagged": 60,
    "issues": 30,
    "spans_word_counts": 30,
    "source_spans_word_counts": 30,
    "spans": 30,
    "source_spans": 30,
    "source_spans_translated": 40,
    "spans_matched": 50,
}

_PRETTIFY_COLS = [
    "spans",
    "source_spans",
    "spans_word_counts",
    "source_spans_word_counts",
    "source_spans_translated",
    "spans_matched",
]


def export_formatted_excel(
    df: pd.DataFrame, path: str, font: str = "Arial", font_size: int = 10,
) -> None:
    """Export DataFrame to Excel with formatting.

    - Frozen header row
    - Top-aligned text in all cells
    - Wider columns for tagged-text columns
    - Hidden columns for raw data (metadata, source_text, label, etc.)
    """
    out = df.copy()
    for col in _PRETTIFY_COLS:
        if col in out.columns:
            out[col] = out[col].apply(_prettify)

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        out.to_excel(writer, index=False)
        ws = writer.sheets["Sheet1"]

        # --- Freeze the header row (row 1) ---
        ws.freeze_panes = "I2"

        # --- Set column widths ---
        for col_idx, col_name in enumerate(out.columns, start=1):
            width = _WIDE_COLS.get(col_name, 10)
            ws.column_dimensions[chr(64 + min(col_idx, 26))].width = width

        # --- Hide specified columns (right-to-left to keep indices stable) ---
        hidden_indices = [i for i, c in enumerate(out.columns) if c in _HIDDEN_COLS]
        for idx in sorted(hidden_indices, reverse=True):
            col_letter = chr(65 + idx)  # A=0, B=1, ...
            ws.column_dimensions[col_letter].hidden = True

        # --- Apply alignment and font to all cells ---
        top_right_align = Alignment(vertical="top", horizontal="left", wrap_text=True)
        font_obj = Font(name=font, size=font_size)
        for row in ws.iter_rows(
            min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column
        ):
            for cell in row:
                cell.alignment = top_right_align
                cell.font = font_obj

        # --- Bold header row (row 1) ---
        for cell in ws[1]:
            cell.font = Font(name=font, size=font_size, bold=True)

    return path


# ---------------------------------------------------------------------------
# Main pipeline
# ---------------------------------------------------------------------------

def main(args: argparse.Namespace) -> None:
    """Run the full validation pipeline."""

    logger.debug("Loading data from %s", args.input_path)
    df = pd.read_json(args.input_path, lines=True)
    logger.info("Loaded %d rows from %s", len(df), args.input_path)

    # # --- Resolve language codes --------------------------------------------
    # src_lang_nllb = _iso2_to_nllb(args.src_lang)
    # tgt_lang_nllb: str | None = None
    # if args.tgt_lang is not None:
    #     tgt_lang_nllb = _iso2_to_nllb(args.tgt_lang)
    #     # print(f"Target language: {args.tgt_lang!r} → {tgt_lang_nllb}")

    # # Cache of already-resolved codes to avoid redundant lookups.
    # _resolved_cache: dict[str, str] = {}

    # def get_tgt_lang(record: dict) -> str:
    #     if tgt_lang_nllb is not None:
    #         return tgt_lang_nllb
    #     raw = record.get(args.lang_col)
    #     if raw is None:
    #         raise KeyError(
    #             f"Row has no field '{args.lang_col}'. "
    #             f"Pass --tgt_lang to use a constant target language instead."
    #         )
    #     if raw not in _resolved_cache:
    #         _resolved_cache[raw] = _iso2_to_nllb(raw)
    #     return _resolved_cache[raw]

    # --- Load model --------------------------------------------------------
    logger.debug("Loading tokenizer from %s", args.tokenizer_path)
    tokenizer = AutoTokenizer.from_pretrained(args.tokenizer_path)
    logger.debug("Loading model from %s", args.model_name_or_path)
    model = AutoModelForSeq2SeqLM.from_pretrained(
        args.model_name_or_path,
        device_map="auto",
        torch_dtype="auto",
        force_download=False,
    )

    # --- Compute heuristics ------------------------------------------------
    logger.info("Computing projection heuristics")
    out = compute_projection_heuristics(df)

    # --- Word counts for original spans ------------------------------------
    logger.debug("Computing word counts for English spans")
    t0 = timer()
    out["spans_word_counts"] = out.apply(
        lambda row: compute_spans_word_counts(row.text, row.label, tokenizer, as_dict=True),
        axis=1,
    )
    logger.debug("  Done in %.1fs", timer() - t0)

    # --- Word counts for projected spans -----------------------------------
    logger.debug("Computing word counts for projected source spans")
    t0 = timer()
    out["source_spans_word_counts"] = out.apply(
        lambda row: compute_spans_word_counts(row.source_text, row.source_label, tokenizer, as_dict=True),
        axis=1,
    )
    logger.debug("  Done in %.1fs", timer() - t0)

    # --- Extract span texts ------------------------------------------------
    out["spans"] = out.apply(
        lambda row: extract_spans(row.text, row.label), axis=1
    )
    out["source_spans"] = out.apply(
        lambda row: extract_spans(row.source_text, row.source_label), axis=1
    )

    # --- Translate spans independently -------------------------------------
    logger.info("Translating projected source spans")
    t0 = timer()
    out["source_spans_translated"] = out.apply(
        lambda row: translate_spans(
            row.source_text, row.source_label, row.lang, # get_tgt_lang(row),
            model, tokenizer, max_length=args.max_length, as_dict=True,
        ),
        axis=1,
    )
    logger.debug("  Done in %.1fs", timer() - t0)

    # --- Free GPU memory before BERTScore ----------------------------------
    model.cpu()
    del model
    torch.cuda.empty_cache()

    # --- BERTScore matching ------------------------------------------------
    logger.info("Computing BERTScore matches (model: %s)", args.bertscore_model)
    matcher = BERTScoreMatcher(model_name=args.bertscore_model)
    t0 = timer()
    out["spans_matched"] = out.apply(
        lambda row: matcher(row.spans, row.source_spans).values.tolist(),
        axis=1,
    )
    logger.debug("  Done in %.1fs", timer() - t0)

    # --- Summary statistics ------------------------------------------------
    n_total = len(out)
    n_issues = (out["n_issues"] > 0).sum()
    logger.debug("\n%s", "=" * 50)
    logger.debug("Summary (%d rows):", n_total)
    logger.debug("  Rows with ≥1 issue : %d (%.1f%%)", n_issues, n_issues / n_total * 100)
    for col in IssueTypes._member_names_:
        n = out[col].sum()
        logger.debug("  %-35s: %d (%.1f%%)", col, n, n / n_total * 100)

    # --- Export ------------------------------------------------------------
    export_formatted_excel(out, args.output_path)
    logger.info("Exported to %s", args.output_path)


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def _build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        description=(
            "Compute projection quality heuristics and export to Excel.  "
            "Reads a JSONL file produced by translate.py (containing projected "
            "span annotations) and flags common projection errors."
        ),
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )

    # --- Logging -------------------------------------------------------------
    log = p.add_argument_group("Logging")
    log.add_argument("--log-level", default="INFO", choices=["DEBUG", "INFO", "WARNING", "ERROR"],
                     help="Root logger level (default: INFO).")
    log.add_argument("-v", "--verbose", action="store_const", dest="log_level", const="DEBUG",
                     help="Shortcut for --log-level DEBUG.")

    # --- I/O ---------------------------------------------------------------
    io = p.add_argument_group("I/O")
    io.add_argument("--input_path", required=True,
                    help="Path to the input JSONL file (output of translate.py).")
    io.add_argument("--output_path", required=True,
                    help="Path to write the formatted Excel (.xlsx) output.")

    # # --- JSONL field names -------------------------------------------------
    # fields = p.add_argument_group("JSONL field names")
    # fields.add_argument("--text_key", default="text",
    #                     help="Key for the source sentence in each JSON object.")
    # fields.add_argument("--template_key", default="template",
    #                     help="Key for the target-language template sentence.")
    # fields.add_argument("--label_key", default="label",
    #                     help="Key for the source-side doccano annotations "
    #                          "(list of [start, end, type]).")
    # fields.add_argument("--output_label_key", default="label_target",
    #                     help="Key under which projected target annotations are "
    #                          "written in the input JSONL.")

    # # --- Language ----------------------------------------------------------
    # lang = p.add_argument_group("Language")
    # lang.add_argument("--src_lang", default="en",
    #                   help="Source language ISO-2 code (e.g. 'en').")
    # lang.add_argument("--tgt_lang", default=None,
    #                   help="Target language ISO-2 code (constant for all rows).  "
    #                        "Mutually exclusive with --lang_col; one is required.")
    # lang.add_argument("--lang_col", default=None,
    #                   help="JSONL field whose value gives the target language per row.  "
    #                        "Used when the target language varies across lines.  "
    #                        "Ignored when --tgt_lang is set.")

    # --- MT Model ----------------------------------------------------------
    model = p.add_argument_group("Translation")
    model.add_argument("--model_name_or_path", default="XIE2021/nllb-200-3.3B-easyproject",
                       help="HuggingFace model ID or local path for the MT model.")
    model.add_argument("--tokenizer_path", default=None,
                       help="HuggingFace tokenizer ID or local path.  "
                            "Defaults to --model_name_or_path.")
    model.add_argument("--max_length", type=int, default=48,
                       help="Maximum sequence length for span translation.")

    # --- BERTScore ---------------------------------------------------------
    bert = p.add_argument_group("BERTScore")
    bert.add_argument("--bertscore_model", default="xlm-roberta-large",
                      help="BERTScore model type (e.g. 'xlm-roberta-large').")

    # --- Excel formatting --------------------------------------------------
    fmt = p.add_argument_group("Excel formatting")
    fmt.add_argument("--font", default="Arial",
                     help="Font name for the Excel output.")
    fmt.add_argument("--font_size", type=int, default=10,
                     help="Font size in points for the Excel output.")

    return p


if __name__ == "__main__":
    parser = _build_parser()
    args = parser.parse_args()

    # Configure logging before anything else.
    numeric_level = getattr(logging, args.log_level.upper(), None)
    if not isinstance(numeric_level, int):
        raise ValueError(f"Invalid log level: {args.log_level}")
    logging.basicConfig(
        format="%(asctime)s [%(levelname)s]: %(message)s",
        datefmt="%H:%M:%S",
        level=numeric_level,
    )

    if args.tokenizer_path is None:
        args.tokenizer_path = args.model_name_or_path
    main(args)
